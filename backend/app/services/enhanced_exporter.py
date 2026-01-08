"""
增强导出服务 - 支持带图表的 Excel/PDF 导出
"""
import io
import json
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd

from app.core.logger import get_logger
from app.services.data_exporter import DataExporter

logger = get_logger(__name__)


class EnhancedExporter:
    """增强导出服务 - 支持多种格式和富内容导出"""
    
    @classmethod
    async def export_with_metadata(
        cls,
        question: str,
        sql: Optional[str],
        columns: List[str],
        rows: List[Dict[str, Any]],
        chart_type: str,
        chart_data: Optional[Dict] = None,
        insight: Optional[str] = None,
        data_interpretation: Optional[Dict] = None,
        fluctuation_analysis: Optional[Dict] = None,
        export_format: str = "excel"
    ) -> Tuple[bytes, str]:
        """
        导出查询结果（包含元数据和分析内容）
        
        Args:
            question: 用户问题
            sql: SQL 查询语句
            columns: 列名列表
            rows: 数据行列表
            chart_type: 图表类型
            chart_data: 图表数据
            insight: AI 洞察
            data_interpretation: 数据解读
            fluctuation_analysis: 波动归因
            export_format: 导出格式（"excel", "excel_with_chart", "pdf", "csv"）
            
        Returns:
            Tuple[bytes, str]: (文件内容, 文件名)
        """
        try:
            if export_format == "csv":
                return cls._export_csv_simple(question, columns, rows)
            elif export_format == "excel":
                return cls._export_excel_with_metadata(
                    question, sql, columns, rows, chart_type,
                    insight, data_interpretation, fluctuation_analysis
                )
            elif export_format == "excel_with_chart":
                return cls._export_excel_with_chart(
                    question, sql, columns, rows, chart_type, chart_data,
                    insight, data_interpretation, fluctuation_analysis
                )
            elif export_format == "pdf":
                return await cls._export_pdf_report(
                    question, sql, columns, rows, chart_type,
                    insight, data_interpretation, fluctuation_analysis
                )
            else:
                raise ValueError(f"不支持的导出格式: {export_format}")
                
        except Exception as e:
            logger.error(f"Enhanced export failed: {e}", exc_info=True)
            # 降级：使用基础导出
            return DataExporter.export_to_excel(rows, columns, question[:20])
    
    @classmethod
    def _export_csv_simple(
        cls,
        question: str,
        columns: List[str],
        rows: List[Dict[str, Any]]
    ) -> Tuple[bytes, str]:
        """导出为简单 CSV"""
        return DataExporter.export_to_csv(rows, columns, question[:20])
    
    @classmethod
    def _export_excel_with_metadata(
        cls,
        question: str,
        sql: Optional[str],
        columns: List[str],
        rows: List[Dict[str, Any]],
        chart_type: str,
        insight: Optional[str],
        data_interpretation: Optional[Dict],
        fluctuation_analysis: Optional[Dict]
    ) -> Tuple[bytes, str]:
        """导出为带元数据的 Excel（多 sheet）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 创建工作簿
            wb = Workbook()
            
            # === Sheet 1: 数据 ===
            ws_data = wb.active
            ws_data.title = "数据"
            
            df = pd.DataFrame(rows, columns=columns)
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 表头样式
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # === Sheet 2: 分析报告 ===
            ws_report = wb.create_sheet(title="分析报告")
            
            # 标题
            ws_report['A1'] = "数据分析报告"
            ws_report['A1'].font = Font(size=16, bold=True)
            ws_report.merge_cells('A1:D1')
            
            row_num = 3
            
            # 问题
            ws_report[f'A{row_num}'] = "问题："
            ws_report[f'A{row_num}'].font = Font(bold=True)
            ws_report[f'B{row_num}'] = question
            ws_report.merge_cells(f'B{row_num}:D{row_num}')
            row_num += 2
            
            # SQL
            if sql:
                ws_report[f'A{row_num}'] = "SQL 查询："
                ws_report[f'A{row_num}'].font = Font(bold=True)
                row_num += 1
                ws_report[f'A{row_num}'] = sql
                ws_report.merge_cells(f'A{row_num}:D{row_num}')
                ws_report[f'A{row_num}'].alignment = Alignment(wrap_text=True)
                row_num += 2
            
            # 数据摘要
            ws_report[f'A{row_num}'] = "数据摘要："
            ws_report[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            ws_report[f'A{row_num}'] = f"共 {len(rows)} 行数据，{len(columns)} 个字段"
            row_num += 1
            ws_report[f'A{row_num}'] = f"图表类型：{chart_type}"
            row_num += 2
            
            # AI 洞察
            if insight:
                ws_report[f'A{row_num}'] = "AI 洞察："
                ws_report[f'A{row_num}'].font = Font(bold=True)
                row_num += 1
                ws_report[f'A{row_num}'] = insight
                ws_report.merge_cells(f'A{row_num}:D{row_num}')
                ws_report[f'A{row_num}'].alignment = Alignment(wrap_text=True)
                row_num += 2
            
            # 数据解读
            if data_interpretation:
                ws_report[f'A{row_num}'] = "数据解读："
                ws_report[f'A{row_num}'].font = Font(bold=True)
                row_num += 1
                
                summary = data_interpretation.get('summary', '')
                ws_report[f'A{row_num}'] = summary
                ws_report.merge_cells(f'A{row_num}:D{row_num}')
                ws_report[f'A{row_num}'].alignment = Alignment(wrap_text=True)
                row_num += 1
                
                key_findings = data_interpretation.get('key_findings', [])
                if key_findings:
                    ws_report[f'A{row_num}'] = "关键发现："
                    ws_report[f'A{row_num}'].font = Font(bold=True, size=11)
                    row_num += 1
                    for finding in key_findings:
                        ws_report[f'A{row_num}'] = f"• {finding}"
                        row_num += 1
                
                row_num += 1
            
            # 波动归因
            if fluctuation_analysis and fluctuation_analysis.get('has_fluctuation'):
                ws_report[f'A{row_num}'] = "波动归因分析："
                ws_report[f'A{row_num}'].font = Font(bold=True)
                row_num += 1
                
                attribution = fluctuation_analysis.get('attribution', {})
                detailed_analysis = attribution.get('detailed_analysis', '')
                if detailed_analysis:
                    ws_report[f'A{row_num}'] = detailed_analysis
                    ws_report.merge_cells(f'A{row_num}:D{row_num}')
                    ws_report[f'A{row_num}'].alignment = Alignment(wrap_text=True)
                    row_num += 1
                
                main_factors = attribution.get('main_factors', [])
                if main_factors:
                    ws_report[f'A{row_num}'] = "主要因素："
                    ws_report[f'A{row_num}'].font = Font(bold=True, size=11)
                    row_num += 1
                    for factor in main_factors:
                        ws_report[f'A{row_num}'] = f"• {factor}"
                        row_num += 1
            
            # 调整列宽
            ws_report.column_dimensions['A'].width = 20
            ws_report.column_dimensions['B'].width = 60
            ws_report.column_dimensions['C'].width = 20
            ws_report.column_dimensions['D'].width = 20
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{question[:20]}_{timestamp}.xlsx"
            
            logger.info(
                "Enhanced Excel export completed",
                rows=len(rows),
                columns=len(columns),
                filename=filename,
                size_kb=len(excel_bytes) / 1024
            )
            
            return excel_bytes, filename
            
        except Exception as e:
            logger.error(f"Enhanced Excel export failed: {e}", exc_info=True)
            # 降级到基础导出
            return DataExporter.export_to_excel(rows, columns, question[:20])
    
    @classmethod
    def _export_excel_with_chart(
        cls,
        question: str,
        sql: Optional[str],
        columns: List[str],
        rows: List[Dict[str, Any]],
        chart_type: str,
        chart_data: Optional[Dict],
        insight: Optional[str],
        data_interpretation: Optional[Dict],
        fluctuation_analysis: Optional[Dict]
    ) -> Tuple[bytes, str]:
        """
        导出为带图表的 Excel（图表嵌入需要前端生成图片后传入）
        
        注意：由于后端难以生成 ECharts 图表图片，此方法暂时与 _export_excel_with_metadata 相同
        实际图表嵌入需要前端先将图表转为图片，然后传给后端
        """
        # TODO: 实现图表嵌入（需要前端传入图表图片）
        logger.info("Chart embedding not yet implemented, falling back to metadata export")
        return cls._export_excel_with_metadata(
            question, sql, columns, rows, chart_type,
            insight, data_interpretation, fluctuation_analysis
        )
    
    @classmethod
    async def _export_pdf_report(
        cls,
        question: str,
        sql: Optional[str],
        columns: List[str],
        rows: List[Dict[str, Any]],
        chart_type: str,
        insight: Optional[str],
        data_interpretation: Optional[Dict],
        fluctuation_analysis: Optional[Dict]
    ) -> Tuple[bytes, str]:
        """导出为 PDF 报告"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # 注册中文字体（如果可用）
            try:
                # 尝试注册常见的中文字体
                import os
                font_paths = [
                    '/System/Library/Fonts/PingFang.ttc',  # macOS
                    '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',  # Linux
                    'C:\\Windows\\Fonts\\msyh.ttc',  # Windows
                ]
                
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                        break
                
                chinese_font = 'ChineseFont'
            except:
                logger.warning("Chinese font registration failed, using default font")
                chinese_font = 'Helvetica'
            
            # 创建 PDF
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            
            # 样式
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=chinese_font,
                fontSize=18,
                spaceAfter=20
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=chinese_font,
                fontSize=14,
                spaceAfter=10
            )
            body_style = ParagraphStyle(
                'CustomBody',
                parent=styles['BodyText'],
                fontName=chinese_font,
                fontSize=10,
                leading=14
            )
            
            # 标题
            story.append(Paragraph("数据分析报告", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # 问题
            story.append(Paragraph(f"<b>问题：</b>{question}", body_style))
            story.append(Spacer(1, 0.1*inch))
            
            # SQL
            if sql:
                story.append(Paragraph("<b>SQL 查询：</b>", body_style))
                story.append(Paragraph(f"<font face='Courier'>{sql[:200]}...</font>", body_style))
                story.append(Spacer(1, 0.1*inch))
            
            # 数据摘要
            story.append(Paragraph(f"<b>数据摘要：</b>共 {len(rows)} 行数据，{len(columns)} 个字段", body_style))
            story.append(Spacer(1, 0.2*inch))
            
            # AI 洞察
            if insight:
                story.append(Paragraph("<b>AI 洞察：</b>", heading_style))
                story.append(Paragraph(insight, body_style))
                story.append(Spacer(1, 0.2*inch))
            
            # 数据解读
            if data_interpretation:
                story.append(Paragraph("<b>数据解读：</b>", heading_style))
                summary = data_interpretation.get('summary', '')
                story.append(Paragraph(summary, body_style))
                story.append(Spacer(1, 0.1*inch))
                
                key_findings = data_interpretation.get('key_findings', [])
                if key_findings:
                    story.append(Paragraph("<b>关键发现：</b>", body_style))
                    for finding in key_findings:
                        story.append(Paragraph(f"• {finding}", body_style))
                
                story.append(Spacer(1, 0.2*inch))
            
            # 波动归因
            if fluctuation_analysis and fluctuation_analysis.get('has_fluctuation'):
                story.append(Paragraph("<b>波动归因分析：</b>", heading_style))
                attribution = fluctuation_analysis.get('attribution', {})
                detailed_analysis = attribution.get('detailed_analysis', '')
                if detailed_analysis:
                    story.append(Paragraph(detailed_analysis, body_style))
                    story.append(Spacer(1, 0.1*inch))
                
                main_factors = attribution.get('main_factors', [])
                if main_factors:
                    story.append(Paragraph("<b>主要因素：</b>", body_style))
                    for factor in main_factors:
                        story.append(Paragraph(f"• {factor}", body_style))
                
                story.append(Spacer(1, 0.2*inch))
            
            # 数据表格（仅显示前 20 行）
            story.append(PageBreak())
            story.append(Paragraph("<b>数据详情（前 20 行）：</b>", heading_style))
            story.append(Spacer(1, 0.1*inch))
            
            # 构建表格数据
            df = pd.DataFrame(rows, columns=columns).head(20)
            table_data = [columns] + df.values.tolist()
            
            # 创建表格
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # 生成 PDF
            doc.build(story)
            pdf_bytes = output.getvalue()
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{question[:20]}_{timestamp}.pdf"
            
            logger.info(
                "PDF export completed",
                rows=len(rows),
                columns=len(columns),
                filename=filename,
                size_kb=len(pdf_bytes) / 1024
            )
            
            return pdf_bytes, filename
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}", exc_info=True)
            # 降级到 Excel 导出
            return cls._export_excel_with_metadata(
                question, sql, columns, rows, chart_type,
                insight, data_interpretation, fluctuation_analysis
            )
